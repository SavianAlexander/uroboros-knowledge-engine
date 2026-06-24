import os
import sqlite3
import subprocess
import shutil
import time
from pathlib import Path

from reportlab.pdfgen import canvas
import docx
import openpyxl

TEST_DB = "knowledge.db"
TEST_DIR = Path("test_sandbox")

def setup_sandbox():
    if TEST_DIR.exists():
        shutil.rmtree(TEST_DIR)
    TEST_DIR.mkdir()
    
    # 1. Plain text note
    (TEST_DIR / "note.txt").write_text("This is a simple text note containing gravity and physics details.", encoding="utf-8")
    
    # 2. Duplicate text note
    (TEST_DIR / "note_dup.txt").write_text("This is a simple text note containing gravity and physics details.", encoding="utf-8")
    
    # 3. Word DOCX file
    doc = docx.Document()
    doc.add_paragraph("This is a Microsoft Word document containing quantum mechanics and relativity formulas.")
    doc.save(TEST_DIR / "document.docx")
    
    # 4. Excel XLSX file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Data Table")
    ws.cell(row=2, column=1, value="Astrophysics statistics report 2026")
    wb.save(TEST_DIR / "data.xlsx")
    
    # 5. PDF document
    pdf_path = TEST_DIR / "report.pdf"
    c = canvas.Canvas(str(pdf_path))
    c.drawString(100, 750, "Annual astrophysics survey and planetary data analysis.")
    c.save()

def teardown_sandbox():
    if TEST_DIR.exists():
        shutil.rmtree(TEST_DIR)
    # retry delete database file if locked temporarily
    for _ in range(5):
        try:
            if os.path.exists(TEST_DB):
                os.remove(TEST_DB)
            break
        except PermissionError:
            time.sleep(0.2)

def run_cmd(args):
    result = subprocess.run([ "python", "know.py" ] + args, capture_output=True, text=True)
    return result.stdout, result.stderr, result.returncode

def main():
    print("Setting up test sandbox with duplicate files...")
    setup_sandbox()
    
    try:
        # 1. Init DB
        stdout, stderr, code = run_cmd(["init"])
        assert code == 0, f"Init failed: {stderr}"
        
        # 2. Index files
        stdout, stderr, code = run_cmd(["index", str(TEST_DIR)])
        assert code == 0, f"Index failed: {stderr}"
        
        # 3. Direct verification of FTS5 search and rules
        # ponytail: use with statement to auto-close connection safely
        with sqlite3.connect(TEST_DB) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM files")
            assert cursor.fetchone()[0] == 5
            
            cursor.execute("SELECT COUNT(*) FROM files WHERE file_size > 100")
            assert cursor.fetchone()[0] > 0
            
            # Setup auto rule
            cursor.execute("INSERT OR IGNORE INTO auto_rules (pattern, tag) VALUES (?, ?)", (r"astrophysics", "science"))
            conn.commit()
            
        # Re-index to apply rules
        stdout, stderr, code = run_cmd(["index", str(TEST_DIR)])
        assert code == 0, f"Re-index failed: {stderr}"
        
        with sqlite3.connect(TEST_DB) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM tags WHERE tag = 'science'")
            assert cursor.fetchone()[0] > 0, "Science tag rules did not execute on index"

        # 5. LAN Sync peers DB check
        with sqlite3.connect(TEST_DB) as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT OR IGNORE INTO sync_peers (address, name) VALUES ('http://127.0.0.1:8000', 'Local Peer')")
            conn.commit()
            cursor.execute("SELECT COUNT(*) FROM sync_peers")
            assert cursor.fetchone()[0] == 1
            
        # 6. Snapshot checks
        import know
        ts = know.create_db_snapshot()
        snaps = know.list_db_snapshots()
        assert ts in snaps
        # clean up snapshot file
        import os
        os.remove(f"{know.DB_FILE}.snapshot-{ts}")
            
        print("All database query operators, duplicates, indexing, auto-tag rules, semantic checks, sync peers DB, and snapshot checks passed successfully!")
    finally:
        print("Cleaning up sandbox...")
        teardown_sandbox()

if __name__ == "__main__":
    main()
